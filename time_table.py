# import
import pandas as pd
import numpy as np

# imports for final excel output
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

# set display options
pd.set_option('display.max_rows', 8)

# list/dictionaries
subject_list = []
period_list = ['TP4', 'TP5', 'TP6', 'TP7']
#period_list = ['HP4','HP5','HP6','HP7']
lesson_lookup = {}
teacher_subject = {}
subject_teacher = {}
faculty_subject = {}
subject_faculty = {}

# read in data frames
teacher = pd.read_csv('C:/Users/sarah/PycharmProjects/portfolio/timetable/teachers.csv', index_col='initials')
SMT = pd.read_csv('C:/Users/sarah/PycharmProjects/portfolio/timetable/smt.csv', index_col='initials')
teacher_timetable = pd.read_csv('C:/Users/sarah/PycharmProjects/portfolio/timetable/tea.csv', index_col='initials')
subject = pd.read_csv('C:/Users/sarah/PycharmProjects/portfolio/timetable/subj.csv', index_col='Subject Code')
room_timetable = pd.read_csv('C:/Users/sarah/PycharmProjects/portfolio/timetable/room.csv', index_col='room')

def lesson_dic(lesson_info):
    class_info = str(lesson_info).split("\n")[0]  # class info without room or teacher (LAC/Cs or 11M/Ma5)

    if len(class_info.split('/')) == 2:  # Check class info can be split

        # input and output yrs
        yr_input = ['5', '6', '7', '8', '9', '10', '11', 'L', 'U']
        yr_output = ['5', '6', '7', '8', '9', '10', '11', '12', '13']

        for i, yr in enumerate(yr_input):
            if class_info.find(yr) == 0:
                yr = yr_output[i]

                # Add lesson to dictionary
                subject = class_info.split('/')[1]
                lesson_lookup[lesson_info] = [yr, subject[:2]]

                # Add set if applicable
                if len(subject) == 3:
                    lesson_lookup[lesson_info].append(f'Set {subject[2]}')
                break


# Teacher timetable remove unnecessary periods and Games and create lesson dictionary
for period in list(teacher_timetable):
    if period not in period_list:
        teacher_timetable.drop(period, axis=1, inplace=True)

    # Remove Games and make lesson dictionary for export
    else:
        for initials in teacher_timetable.index.values:
            lesson_info = teacher_timetable.at[initials, period]
            class_info = str(lesson_info).split("\n")[0] # class info without room (LAC/Cs or 11M/Ma5)

            if 'Games' in class_info: # Remove Games and Off Games
                teacher_timetable.at[initials, period]= np.NaN
            elif 'Meeting' in class_info: # Remove Games and Off Games
                teacher_timetable.at[initials, period]= np.NaN
            elif 'Part Time' in class_info: # Remove Games and Off Games
                teacher_timetable.at[initials, period]= np.NaN

            lesson_dic(lesson_info)

            if lesson_info not in lesson_lookup.keys(): # final check for annomalies in data
                teacher_timetable.at[initials, period] = np.NaN

# create list of teachers for cover
for initials in teacher.index.values:
    if initials not in SMT.index.values:
        subject_list.append(initials)

# create dictionary teachers name to subjects taught
for initials in teacher_timetable.index.values: # Add all teachers first
    teacher_subject[initials]=[]

for subject_abr in subject.index.values:
    if subject_abr not in ['Ga', 'Tk', 'Ps', 'Ls']:
        subject_teacher[subject_abr] = []
        for initials in list(subject)[3:]:

            staff_initials = subject.at[subject_abr, initials]

            if staff_initials is not np.NaN and pd.notna(staff_initials): # remove na
                if str(staff_initials).find(', ') == -1:  # check for multiple entries (applicable in HoDs column)
                    if staff_initials in teacher_subject.keys() and subject_abr not in teacher_subject[staff_initials]:
                        teacher_subject[staff_initials].append(subject_abr)
                    if staff_initials not in subject_teacher[subject_abr]:
                        subject_teacher[subject_abr].append(staff_initials)
                else:
                    for i in staff_initials.split(', '):
                        if i in teacher_subject.keys() and subject_abr not in teacher_subject[i]:
                            teacher_subject[i].append(subject_abr)

# Reverse each list as HoDs need to be last to be picked
for staff_initials in subject_teacher.keys():
    subject_teacher[staff_initials].reverse()

# Room timetable remove periods and games
for period in list(room_timetable):
    if period not in period_list:
        room_timetable.drop(period, axis=1, inplace=True)
    else:
        for initials in room_timetable.index.values:
            lesson_info = room_timetable.at[initials, period]
            class_info = str(lesson_info).split("\n")[0]  # class info without room (LAC/Cs or 11M/Ma5)

            if 'Games' in class_info:  # Remove Games and Off Games
                room_timetable.at[initials, period] = 'NaN'

            lesson_dic(lesson_info)

            if lesson_info not in lesson_lookup.keys(): # final check for annomalies in data
                room_timetable.at[initials, period] = np.NaN


# Define faculties to subject dictionaries
faculty_subject["Languages"] = ["EnL", "En", "Sp", "Fr", "Gm", "Gr", "In", "La", "Ru", "Cl", "La", "Ja", "Cn"]
faculty_subject["Business/Economics"] = ["Ec", "Bs", "Bm", "BsB"]
faculty_subject["Philosophy/Psychology"] = ["Py", "Pp", "Po", "So", "Ps", "Tk", "Dp"]
faculty_subject["Geography/History"] = ["Hi", "Gg", "Gp"]
faculty_subject["Science/Maths"] = ["Ma", "Ph", "Ch", "Bi", "Dt", "As", "Cs", "IT", "Sc"]
faculty_subject["Arts"]=["Ar", "Dr", "Mu"]

for faculty in faculty_subject.keys():
    for subject_abr in faculty_subject[faculty]:
        if subject_abr not in subject_faculty.keys():
            subject_faculty[subject_abr]=faculty

# free teachers/room for a period of interest, period: 0 = 4/6, 1 = 5/7
def free(teacher, period):
    free = []
    clash = []
    for teacher_initials in teacher.index.values:
        first_lesson = teacher.at[teacher_initials, period_list[period]]
        second_lesson = teacher.at[teacher_initials, period_list[period + 2]]
        if first_lesson is np.NaN and second_lesson is np.NaN:
            free.append(teacher_initials)
        elif first_lesson is not np.NaN and second_lesson is not np.NaN:

            clash.append([teacher_initials,first_lesson,second_lesson])
    return(free,clash)


# return both free and clashing teachers and staff across both pairs of periods, named 0 and 1
free_teacher_0, clash_teacher_0 = free(teacher_timetable, 0)
free_room_0, clash_room_0 = free(room_timetable, 0)
free_teacher_1, clash_teacher_1 = free(teacher_timetable, 1)
free_room_1, clash_room_1 = free(room_timetable, 1)


def teacher_collapse(free, clash):
    resolution = {}
    for clash_info in clash:
        teacher = clash_info[0]
        first_lesson = clash_info[1] # stay and cover
        second_lesson = clash_info[2]  # move with teacher

        subject = lesson_lookup[first_lesson][1]

        for cover_teacher in subject_teacher[subject]: # for pot cover teacher within subject teachers
            if cover_teacher in free: # if the teacher is free
                resolution[teacher]=[cover_teacher,first_lesson,second_lesson] # assign cover
                free.remove(cover_teacher)
                break
    return resolution, free


def room_collapse(free, clash):
    resolution = {}
    for clash_info in clash:
        room = clash_info[0]
        first_lesson = clash_info[1] # stay and cover
        second_lesson = clash_info[2]  # move with teacher

        resolution[room]=[free[0], first_lesson, second_lesson]
        free.remove(free[0])

    return resolution, free


teacher_res_0, teacher_free_0 = teacher_collapse(free_teacher_0, clash_teacher_0)
teacher_res_1, teacher_free_1 = teacher_collapse(free_teacher_1, clash_teacher_1)
room_res_0, room_free_0 = room_collapse(free_room_0, clash_room_0)
room_res_1, room_free_1 = room_collapse(free_room_1, clash_room_1)

# Export
wb = Workbook()
file = 'C:/Users/sarah/PycharmProjects/portfolio/timetable/tt_collapse.xlsx'
wb.save(file)

def exp_t(file, title, period, res, free):
    wb = load_workbook(file)
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws.column_dimensions["A"].width = 5
    ws.cell(2, 2).value = "Teacher"
    ws.column_dimensions["B"].width = 10
    ws.cell(2, 3).value = "P" + str(period) + " Lesson"
    ws.column_dimensions["C"].width = 10
    ws.cell(2, 4).value = "P" + str(period + 2) + " Lesson"
    ws.column_dimensions["D"].width = 10
    ws.cell(2, 5).value = "Resolution"
    ws.column_dimensions["E"].width = 30
    ws.cell(2, 6).value = "Cover Teacher"
    ws.column_dimensions["F"].width = 10
    ws.cell(2, 8).value = "Free Teachers"
    ws.column_dimensions["H"].width = 10

    for r,teacher in enumerate(res.keys()):
        first_lesson = lesson_lookup[res[teacher][1]]
        first_room = res[teacher][1].split('\n')
        if len(first_room) == 2:
            first_room = first_room[1]
        else:
            first_room = 'TBC'
        second_lesson = lesson_lookup[res[teacher][2]]

        ws.cell(3+r,2).value = teacher
        ws.cell(3+r,3).value = f'Yr{first_lesson[0]} - {first_lesson[1]}'
        ws.cell(3 + r,4).value = f'Yr{second_lesson[0]} - {second_lesson[1]}'
        ws.cell(3 + r, 5).value = f'Yr{second_lesson[0]} - {second_lesson[1]} to be covered in {first_room}'
        ws.cell(3 + r, 6).value = res[teacher][0]

    for r,teacher in enumerate(free):
        ws.cell(3+r, 8).value = teacher

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    for row in range(2,ws.max_row+1):
        for column in [2,3,4,5,6,8]:
            ws.cell(row=row, column=column).border = thin_border

    wb.save(file)


exp_t(file, 'Teacher 4-6', 4, teacher_res_0, teacher_free_0)
exp_t(file, 'Teacher 5-7', 5, teacher_res_1, teacher_free_1)

def exp_r(file, title, period, res, free):
    wb = load_workbook(file)
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws.column_dimensions["A"].width = 5
    ws.cell(2, 2).value = "Room"
    ws.column_dimensions["B"].width = 10
    ws.cell(2, 3).value = "P" + str(period) + " Lesson"
    ws.column_dimensions["C"].width = 10
    ws.cell(2, 4).value = "P" + str(period + 2) + " Lesson"
    ws.column_dimensions["D"].width = 10
    ws.cell(2, 5).value = "Resolution"
    ws.column_dimensions["E"].width = 30
    ws.cell(2, 6).value = "New Room"
    ws.column_dimensions["F"].width = 10
    ws.cell(2, 8).value = "Free Room"
    ws.column_dimensions["H"].width = 10

    for r,teacher in enumerate(res.keys()):
        first_lesson = lesson_lookup[res[teacher][1]]
        first_teacher = res[teacher][1].split('\n')[1]
        second_lesson = lesson_lookup[res[teacher][2]]

        ws.cell(3+r,2).value = teacher
        ws.cell(3+r,3).value = f'Yr{first_lesson[0]} - {first_lesson[1]}'
        ws.cell(3 + r,4).value = f'Yr{second_lesson[0]} - {second_lesson[1]}'
        ws.cell(3 + r, 5).value = f'Yr{first_lesson[0]} - {first_lesson[1]} with {first_teacher} to move rooms'
        ws.cell(3 + r, 6).value = res[teacher][0]

    for r,teacher in enumerate(free):
        ws.cell(3+r, 8).value = teacher

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    for row in range(2,ws.max_row+1):
        for column in [2,3,4,5,6,8]:
            ws.cell(row=row, column=column).border = thin_border

    wb.save(file)


exp_r(file, 'Room 4-6', 4, room_res_0, room_free_0)
exp_r(file, 'Room 5-7', 5, room_res_1, room_free_1)

wb = load_workbook(file)
ws = wb["Sheet"]
wb.remove(ws)
wb.save(file)
